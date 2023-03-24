import os
from PyPDF2 import PdfReader
import zipfile
from openpyxl import load_workbook

"""
Написать автотест, запаковывающий в zip-архив несколько разных файлов – pdf, xlsx, csv,
и положить его в папку resources. Реализовать чтение и проверку в ассертах содержимого
каждого файла из архива, не распаковывая его.
"""

res_path = os.path.abspath("resources")
csv_path = os.path.join(res_path, "SampleCSVFile_2kb.csv")
xlsx_path = os.path.join(res_path, "file_example_XLSX_50.xlsx")
pdf_path = os.path.join(res_path, "file-sample_150kB.pdf")


def test_pdf(zip_files):
    reader_origin = PdfReader(pdf_path)
    page_count = len(reader_origin.pages)

    with zipfile.ZipFile('resources/archive.zip') as zf:
        with zf.open('file-sample_150kB.pdf', 'r') as pdf_read:
            reader_arch = PdfReader(pdf_read)
            page_count_arch = len(reader_arch.pages)
            text = reader_arch.pages[0].extract_text()
    assert page_count == page_count_arch
    assert "Lorem ipsum" in text


def test_csv(zip_files):
    with open(csv_path, 'r') as csv_origin:
        row_count = len(list(csv_origin))

    with zipfile.ZipFile('resources/archive.zip') as zf:
        with zf.open('SampleCSVFile_2kb.csv', 'r') as csv_file:
            row_count_archive = len(list(csv_file))

    assert row_count == row_count_archive


def test_xlsx(zip_files):
    wb = load_workbook(xlsx_path)
    sheet = wb.active
    row_count_origin = sheet.max_row
    column_count_origin = sheet.max_column

    with zipfile.ZipFile('resources/archive.zip') as zf:
        with zf.open('file_example_XLSX_50.xlsx', 'r') as xlsx_file:
            wb_arch = load_workbook(xlsx_file)
            sheet_arch = wb_arch.active
            row_count_arch = sheet_arch.max_row
            column_count_arch = sheet_arch.max_column

    assert row_count_origin == row_count_arch
    assert column_count_origin == column_count_arch
